
import sys# Add parent directory to Python path for imports
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from parsing_operations import parse_schedule_query
from typing import Dict, List
import pandas as pd
import openpyxl
import unittest
import json
import azure.functions as func
from tabulate import tabulate



class MockHttpRequest:
    def __init__(self, body=None):
        self._body = body if body is not None else b''

    def get_body(self):
        return self._body

class TestResultRecorder:
    def __init__(self):
        self.results: List[Dict] = []

    def record_result(self, request_input: str, response_output: str):
        self.results.append({
            "Request Input": request_input,
            "Response Output": response_output
        })

    def print_results_table(self):
        print("\nTest Results:")
        print(tabulate(self.results, headers="keys", tablefmt="pipe"))

    def save_to_excel(self, filename="test_results.xlsx"):
        """Save test results to an Excel file with formatted table and word wrapping"""
        df = pd.DataFrame(self.results)
        
        # Ensure the output directory exists
        output_dir = os.path.join(os.path.dirname(__file__), 'test-reports')
        os.makedirs(output_dir, exist_ok=True)
        
        # Save to Excel
        excel_path = os.path.join(output_dir, filename)
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Test Results')
            
            workbook = writer.book
            worksheet = writer.sheets['Test Results']
            
            # Format headers
            header_fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            header_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Format data cells
            data_alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
            thin_border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )
            
            # Auto-adjust column widths and format cells
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).str.len().max(),
                    len(col)
                )
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_length + 2, 50)
                
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=idx + 1)
                    cell.alignment = data_alignment
                    cell.border = thin_border
            
            # Auto-adjust row heights
            for row in worksheet.rows:
                max_height = 0
                for cell in row:
                    if cell.value:
                        text_lines = str(cell.value).count('\n') + 1
                        max_height = max(max_height, text_lines * 15)
                if max_height > 0:
                    worksheet.row_dimensions[cell.row].height = max_height
        
        print(f"\nTest results saved to: {excel_path}")

class TestParseScheduleQuery(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.recorder = TestResultRecorder()

    def _execute_test(self, request_body, expected_status_code: int, 
                     expected_response_status: str, expected_message: str = None,
                     test_description: str = None):
        """Helper method to execute common test steps"""
        if isinstance(request_body, str):
            request_body = request_body.encode('utf-8')
        mock_request = MockHttpRequest(request_body)
        
        response = parse_schedule_query(mock_request)
        
        self.assertEqual(response.status_code, expected_status_code)
        response_body = json.loads(response.get_body())
        self.assertEqual(response_body["status"], expected_response_status)
        
        if expected_message:
            if isinstance(expected_message, str):
                self.assertEqual(response_body.get("message", response_body.get("UserQuery")), expected_message)
            else:  # for regex or partial matching
                self.assertTrue(expected_message(response_body.get("message", "")))
        
        self.recorder.record_result(
            test_description or str(request_body),
            f"{response.get_body().decode('utf-8')} (Status: {response.status_code})"
        )
        
        return response_body

    def test_valid_queries(self):
        """Test various valid user queries"""
        test_cases = [
            "Schedule a meeting with John tomorrow at 2pm",
            "Find available time slots for team meeting next week",
            "Book a room for project review on Friday",
            "Set up a 1-hour meeting with marketing team",
            "Schedule daily standup at 10am starting next Monday"
        ]
        
        for query in test_cases:
            with self.subTest(query=query):
                response_body = self._execute_test(
                    request_body=query,
                    expected_status_code=200,
                    expected_response_status="success",
                    expected_message=query,
                    test_description=f"Valid query: '{query}'"
                )
                self.assertEqual(response_body["UserQuery"], query)

    def test_special_characters(self):
        """Test queries with special characters"""
        test_cases = [
            "Meeting with @john.doe & @jane.smith",
            "Project sync-up: Q1'24 planning",
            "Review meeting (high-priority!)",
            "Team lunch @ Café Noir",
            "Status update: 100% completion"
        ]
        
        for query in test_cases:
            with self.subTest(query=query):
                response_body = self._execute_test(
                    request_body=query,
                    expected_status_code=200,
                    expected_response_status="success",
                    expected_message=query,
                    test_description=f"Special characters: '{query}'"
                )

    def test_escape_characters(self):
        """Test queries with escape characters"""
        test_cases = [
            "Meeting\nwith\tJohn",
            "Project\\Review\\Session",
            "Team\r\nSync",
            "Status: \"In Progress\"",
            "Notes: 'Important\\'s meeting'"
        ]
        
        for query in test_cases:
            with self.subTest(query=query):
                response_body = self._execute_test(
                    request_body=query,
                    expected_status_code=200,
                    expected_response_status="success",
                    expected_message=query,
                    test_description=f"Escape characters: '{query}'"
                )

    def test_empty_request_body(self):
        """Test with an empty request body"""
        self._execute_test(
            request_body=b'',
            expected_status_code=400,
            expected_response_status="error",
            expected_message="Empty request body",
            test_description="Empty request body"
        )

    def test_none_request_body(self):
        """Test with None as request body"""
        self._execute_test(
            request_body=None,
            expected_status_code=400,
            expected_response_status="error",
            expected_message="Empty request body",
            test_description="None request body"
        )

    def test_invalid_encoding(self):
        """Test with invalid UTF-8 encoding"""
        self._execute_test(
            request_body=b'\xff\xfe\xfd',
            expected_status_code=500,
            expected_response_status="error",
            expected_message=lambda msg: "Error processing request" in msg,
            test_description="Invalid UTF-8 encoding"
        )

    def test_extremely_long_query(self):
        """Test with an extremely long query"""
        long_query = "Schedule a meeting " * 1000
        self._execute_test(
            request_body=long_query,
            expected_status_code=200,
            expected_response_status="success",
            expected_message=long_query,
            test_description="Extremely long query (truncated): " + long_query[:100] + "..."
        )

    @classmethod
    def tearDownClass(cls):
        cls.recorder.print_results_table()
        cls.recorder.save_to_excel()

def run_tests():
    """Execute the test suite and display results"""
    suite = unittest.TestLoader().loadTestsFromTestCase(TestParseScheduleQuery)
    runner = unittest.TextTestRunner(verbosity=2)
    runner.run(suite)

if __name__ == '__main__':
    run_tests() 