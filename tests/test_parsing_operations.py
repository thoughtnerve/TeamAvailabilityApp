import unittest
import json
import azure.functions as func
from tabulate import tabulate
import sys
import os

# Add parent directory to Python path for imports
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from parsing_operations import parse_schedule_query
from typing import Dict, List
import pandas as pd
import openpyxl

class MockHttpRequest:
    def __init__(self, body=None):
        self._body = body if body is not None else b''

    def get_body(self):
        return self._body

class TestResultRecorder:
    def __init__(self):
        self.results: List[Dict] = []

    def record_result(self, request_input: str, response_output: str):
        self.results.append({
            "Request Input": request_input,
            "Response Output": response_output
        })

    def print_results_table(self):
        print("\nTest Results:")
        print(tabulate(self.results, headers="keys", tablefmt="pipe"))

    def save_to_excel(self, filename="test_results.xlsx"):
        """Save test results to an Excel file with formatted table and word wrapping
        
        Args:
            filename (str): Name of the Excel file to create
        """
        # Create a pandas DataFrame from the results
        df = pd.DataFrame(self.results)
        
        # Ensure the output directory exists
        output_dir = os.path.join(os.path.dirname(__file__), 'test-reports')
        os.makedirs(output_dir, exist_ok=True)
        
        # Save to Excel
        excel_path = os.path.join(output_dir, filename)
        
        # Create Excel writer object
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Write DataFrame to Excel
            df.to_excel(writer, index=False, sheet_name='Test Results')
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Test Results']
            
            # Get the dimensions of the data
            max_row = len(df) + 1  # +1 for header
            max_col = len(df.columns)
            
            # Create a table
            tab = openpyxl.worksheet.table.Table(
                displayName="TestResults",
                ref=f"A1:{chr(64 + max_col)}{max_row}"
            )
            
            # Add a style to the table
            style = openpyxl.worksheet.table.TableStyleInfo(
                name="TableStyleMedium9",  # This style has a dark header
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            tab.tableStyleInfo = style
            worksheet.add_table(tab)
            
            # Format headers (row 1)
            header_fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            header_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Format data cells
            data_alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
            thin_border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )
            
            # Format column widths and data cells
            for idx, col in enumerate(df.columns):
                column_width = max(
                    df[col].astype(str).str.len().max(),  # max length of values
                    len(col)  # length of column name
                )
                # Set column width (max 50 characters)
                worksheet.column_dimensions[chr(65 + idx)].width = min(column_width + 2, 50)
                
                # Format data cells (skip header row)
                for row in range(2, max_row + 1):
                    cell = worksheet.cell(row=row, column=idx + 1)
                    cell.alignment = data_alignment
                    cell.border = thin_border
            
            # Auto-adjust row heights based on content
            for row in worksheet.rows:
                max_height = 0
                for cell in row:
                    if cell.value:
                        text_lines = str(cell.value).count('\n') + 1
                        max_height = max(max_height, text_lines * 15)  # 15 points per line
                if max_height > 0:
                    worksheet.row_dimensions[cell.row].height = max_height
        
        print(f"\nTest results saved to: {excel_path}")

class TestParseScheduleQuery(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.recorder = TestResultRecorder()

    def _execute_test(self, request_body, expected_status_code: int, 
                     expected_response_status: str, expected_message: str = None,
                     test_description: str = None):
        """Helper method to execute common test steps
        
        Args:
            request_body: The body to use in the mock request
            expected_status_code: Expected HTTP status code
            expected_response_status: Expected status in response body ("success" or "error")
            expected_message: Expected message in response body (optional)
            test_description: Description of the test for the results table
        """
        if isinstance(request_body, str):
            request_body = request_body.encode('utf-8')
        mock_request = MockHttpRequest(request_body)
        
        response = parse_schedule_query(mock_request)
        
        self.assertEqual(response.status_code, expected_status_code)
        response_body = json.loads(response.get_body())
        self.assertEqual(response_body["status"], expected_response_status)
        
        if expected_message:
            if isinstance(expected_message, str):
                self.assertEqual(response_body["message"], expected_message)
            else:  # for regex or partial matching
                self.assertTrue(expected_message(response_body["message"]))
        
        self.recorder.record_result(
            test_description or str(request_body),
            f"{response.get_body().decode('utf-8')} (Status: {response.status_code})"
        )
        
        return response_body

    def test_valid_query(self):
        """Test with a valid user query"""
        test_query = "Schedule a meeting with John tomorrow at 2pm"
        response_body = self._execute_test(
            request_body=test_query,
            expected_status_code=200,
            expected_response_status="success",
            test_description=f"Valid text query: '{test_query}'"
        )
        self.assertEqual(response_body["UserQuery"], test_query)

    def test_empty_request_body(self):
        """Test with an empty request body"""
        self._execute_test(
            request_body=b'',
            expected_status_code=400,
            expected_response_status="error",
            expected_message="Empty request body",
            test_description="Empty request body"
        )

    def test_none_request_body(self):
        """Test with None as request body"""
        self._execute_test(
            request_body=None,
            expected_status_code=400,
            expected_response_status="error",
            expected_message="Empty request body",
            test_description="None request body"
        )

    def test_invalid_encoding(self):
        """Test with invalid UTF-8 encoding"""
        self._execute_test(
            request_body=b'\xff\xfe\xfd',
            expected_status_code=500,
            expected_response_status="error",
            expected_message=lambda msg: "Error processing request" in msg,
            test_description="Invalid UTF-8 encoding"
        )

    @classmethod
    def tearDownClass(cls):
        cls.recorder.print_results_table()
        cls.recorder.save_to_excel()


def run_tests():
    """Execute the test suite and display results"""
    # Create a test suite
    suite = unittest.TestLoader().loadTestsFromTestCase(TestParseScheduleQuery)
    
    # Run tests with console output
    runner = unittest.TextTestRunner(verbosity=2)
    runner.run(suite)

if __name__ == '__main__':
    run_tests()
